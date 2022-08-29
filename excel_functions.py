import openpyxl
import datetime
import os
import re
import json
from openpyxl.styles import Font, PatternFill, Side, Border, Alignment

def check_file_path(file_path: str):
    try:
        f = open(file_path, 'r')
        f.close()
    except IOError:
        return False
    return True

def check_folder_path(folder_path: str) -> bool:
    """Creates folder if it does not exist."""
    if os.path.exists(folder_path) and os.path.isdir(folder_path):
        return True
    else:
        try:
            os.mkdir(folder_path)
            return check_folder_path(folder_path)
        except WindowsError:
            return False


def check_file_name(name: str):
    pattern = r'.*_\d{2}_\d{2}.{1}xlsx'
    return True if not name or re.findall(pattern, name) else False

def make_file_name(name: str, extension: str, substr: bool = False):
    """Returns the file name with specified extension as a string """
    date = f'_{datetime.datetime.now().strftime("%d_%m")}'
    site = ''
    if name.endswith(extension):
        name = name.strip().rsplit(extension, 1)[0]
    if substr:
        site = '_{site}'
        # url = f'_{url.lstrip("https://").split("/")[0].replace(".", "_")}'
    return f'{name}{site}{date}{extension}'

def del_dir_files(dir_path: str) -> bool:
    """Returns True if dir is empty."""
    for file in os.listdir(dir_path):
        os.remove(os.path.join(dir_path, file))
    if os.listdir(dir_path):
        return del_dir_files(dir_path)
    else:
        return True

def get_worksheet(file_path: str):
    wb = openpyxl.load_workbook(file_path)
    sh = wb.active
    title_row = list(sh[1])
    return sh, title_row

def get_table(w_sheet, min_row: int, max_col: int):
    main_table = []
    max_row = 0
    for i, row in enumerate(w_sheet.iter_rows(min_row=min_row, max_col=max_col)):
        new_row = []
        for cell in row:
            if cell.value is not None:
                new_row.append(cell.value)
            else:
                new_row.append('')
        max_row = i + min_row + 1
        if new_row == [''] * len(new_row):
            break
        main_table.append(new_row)

    return main_table, max_row

def get_file_from_data(folder_path: str, file_name: str, data: list, columns_names: list, styles: list):
    """
    Save data to the file with 'xlsx' extension.
    :param folder_path: folder_path
    :param file_name: file_name
    :param data: list of lists, each of which will be a row in file
    :param columns_names: names of columns in file.
    :param styles: list of styles colors.
    :return: error_msg
    """
    full_path = os.path.join(folder_path, file_name)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Main'
    error_msg = ''
    if len(columns_names) != len(max(data, key=len, default=[])):
        error_msg = f'Количество имён колонок не соответствует количеству столбцов данных.\n'
    else:
        for row in [columns_names] + data:
            if isinstance(row[0], (str, int)):
                ws.append(row)

        thin = Side(border_style='thin', color='000000')
        for i, row in enumerate(ws.iter_rows()):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if i:
                    cell.fill = PatternFill('solid', fgColor='FFFFFF')
                    cell.font = Font(color='000000')
                else:
                    cell.fill = PatternFill('solid', fgColor=styles[0])
                    cell.font = Font(color=styles[1])

        try:
            wb.save(full_path)
        except PermissionError:
            error_msg = f'Ошибка доступа к файлу {full_path}. Закройте файл.\n'
    return error_msg

def get_image_fields(alias: str, images_paths: list) -> tuple:
    """
    Returns list of tuples, if length of paths smaller than 12 list will be added elements like empty string
    :param alias:
    :param images_paths:
    :return:
    """
    num = 13
    if len(images_paths) > num:
        images_paths = images_paths[: num - 1]
    site_path = "assets/images/product_images/"
    file_names = [f'{alias}.jpg'] + [f'{alias}_{str(i)}.jpg' for i in range(1, len(images_paths))]
    values = [site_path + f_name for f_name in file_names]
    if len(values) < num:
        values += [''] * (num - len(values))
    fields_names = ['image'] + [f'image_{str(i)}' for i in range(1, num)]
    return images_paths, file_names, values, fields_names

def check_version(version: str):
    json_str = {
                "uploader": {"version": version},
                "save_dir_path": {"path": ""},
                "db_file_name": {"name": ""},
                "file_name": {"name": ""},
                "images_folder_name": {"name": ""},
                "title_fill_color": {"color": ""},
                "title_font_color": {"color": ""},
                "update_file_name": {"name": ""},
                "del_file_name": {"name": ""},
                "columns": {"names": []},
                "category_site_file_name": {"name": ""},
                "data_site_file_name": {"name": ""},
                "category_site_urls": {"urls": {}},
                "functions_names": {"names": ["get_categories",
                                              "get_products",
                                              "get_item_content",
                                              "get_item_images"]}
                }

    version_ = ''
    if check_file_path('settings.json'):
        with open('settings.json') as json_file:
            data = json.load(json_file)
            if 'uploader' in data.keys():
                version_ = data['uploader']['version']

    if not check_file_path('settings.json') or version_ != version:
        with open('settings.json', 'w', encoding='utf-8',) as file:
            file.write(json.dumps(json_str))
